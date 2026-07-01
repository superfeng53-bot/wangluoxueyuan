"""
Excel 导入/导出（中文表头，openpyxl）。
复制到 <svc>/web/excel_io.py，按 site_profile 调整：
  - A 型：IMPORT_COLS 保持默认（含学科/学分/卡号）
  - B 型：替换 IMPORT_COLS 为 B_IMPORT_COLS，删除 A 型学科列

依赖：openpyxl（已在项目 requirements.txt 中）
"""
from __future__ import annotations

import io
import json
from dataclasses import dataclass
from typing import Any

import openpyxl

try:
    from ..credential_parser import CredentialParseError, parse_combined_credentials
except ImportError:
    from .credential_parser import CredentialParseError, parse_combined_credentials
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── A 型列定义（含学科/学分/卡号）────────────────────────────────────────────

IMPORT_COLS = [
    "姓名", "账号", "密码",
    "学科1", "学分1", "学科2", "学分2",
    "卡号", "卡号密码", "备注",
]

# 一栏凭证列（combined 模式插入在姓名/账号之前）
COMBINED_CRED_COL = "账号密码"

A_IMPORT_COLS_COMBINED = [
    "姓名", COMBINED_CRED_COL, "账号", "密码",
    "学科1", "学分1", "学科2", "学分2",
    "卡号", "卡号密码", "备注",
]

# ── B 型列定义（公需年度型）─ 替换上面的 IMPORT_COLS ──────────────────────────

B_IMPORT_COLS = [
    "账号", "密码", "备注", "目标年度", "任务模式",
]

B_IMPORT_COLS_COMBINED = [
    COMBINED_CRED_COL, "账号", "密码", "备注", "目标年度", "任务模式",
]

# ── 导出追加列（状态/日志等）─────────────────────────────────────────────────

EXPORT_EXTRA_COLS = [
    "状态", "说明", "重试次数", "创建时间", "更新时间",
    "最近运行结果", "错误日志",
]

# B 型导出：登录后字段 + 系统列（excel-spec.md §2B）
B_EXPORT_EXTRA_COLS = [
    "姓名", "身份证", "状态", "说明", "重试次数", "创建时间", "更新时间",
    "最近运行结果", "错误日志",
]

B_YEAR_HEADER_ALIASES = ("目标年度", "年度", "年份", "target_years")
B_MODE_HEADER_ALIASES = ("任务模式", "上报模式", "report_mode")

PLATFORM_NAME = "成都职业培训网络学院"


def _current_year_str() -> str:
    from datetime import datetime
    from zoneinfo import ZoneInfo
    return str(datetime.now(tz=ZoneInfo("Asia/Shanghai")).year)


def import_cols_for(site_profile: str = "A", credential_input_mode: str = "split") -> list[str]:
    """按 site_profile 与凭证输入模式返回导入/导出前半段列顺序。"""
    profile = site_profile.upper()
    combined = (credential_input_mode or "split").strip().lower() == "combined"
    if profile == "B":
        return B_IMPORT_COLS_COMBINED if combined else B_IMPORT_COLS
    return A_IMPORT_COLS_COMBINED if combined else IMPORT_COLS


def format_combined_credential(username: str, password: str) -> str:
    """导出用：生成可再次导入解析的一栏凭证文本。"""
    user = (username or "").strip()
    pwd = (password or "").strip()
    if not user:
        return ""
    if pwd:
        return f"账号 {user} 密码 {pwd}"
    return f"账号 {user}"


def resolve_credentials_from_row(
    username: str,
    password: str,
    combined: str = "",
) -> tuple[str, str]:
    """从分列或一栏列解析账号密码；分列均有值时优先分列。"""
    user = (username or "").strip()
    pwd = (password or "").strip()
    comb = (combined or "").strip()
    if user and pwd:
        return user, pwd
    if comb:
        parsed = parse_combined_credentials(comb)
        return parsed.username, parsed.password
    if user or pwd:
        raise CredentialParseError("账号和密码须同时填写，或改填「账号密码」一栏")
    raise CredentialParseError("账号和密码不能均为空")


def _resolve_header_col(header: list[str], names: tuple[str, ...]) -> int | None:
    for name in names:
        try:
            return header.index(name)
        except ValueError:
            continue
    return None


# ── 样式 ──────────────────────────────────────────────────────────────────────

def _header_style() -> dict:
    return {
        "font": Font(name="微软雅黑", bold=True, size=11),
        "fill": PatternFill("solid", fgColor="3659F0"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="thin", color="FFFFFF"),
        ),
    }


def _apply_style(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


# ── 模板生成 ──────────────────────────────────────────────────────────────────

def build_template_xlsx(
    import_cols: list[str] | None = None,
    site_profile: str = "A",
    credential_input_mode: str = "split",
) -> bytes:
    """生成 {PLATFORM_NAME}账号模板.xlsx，含账号列表 + 填写说明两个 Sheet。"""
    if import_cols is None:
        cols = import_cols_for(site_profile, credential_input_mode)
    else:
        cols = import_cols
    wb = openpyxl.Workbook()

    # Sheet1：账号列表
    ws1 = wb.active
    ws1.title = "账号列表"
    ws1.row_dimensions[1].height = 28
    style = _header_style()
    for ci, col_name in enumerate(cols, start=1):
        cell = ws1.cell(row=1, column=ci, value=col_name)
        cell.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
        ws1.column_dimensions[get_column_letter(ci)].width = max(12, len(col_name) * 2 + 4)

    # 示例行
    sample = _get_sample_row(cols)
    for ci, val in enumerate(sample, start=1):
        cell = ws1.cell(row=2, column=ci, value=val)
        cell.font = Font(name="微软雅黑", size=10, color="888888")

    # Sheet2：填写说明
    ws2 = wb.create_sheet("填写说明")
    combined = COMBINED_CRED_COL in cols
    if site_profile.upper() == "B":
        notes = [
            ("说明", "内容"),
            ("表头", "不可改字、不可调列顺序"),
            ("必填列", "账号+密码（分列填写），或仅填「账号密码」一栏（自动识别）"),
            *(
                [("账号密码", "可粘贴「账号 xxx 密码 yyy」或 user pass；与分列二选一")]
                if combined else []
            ),
            ("目标年度", "可多选，用顿号或逗号分隔；不填则默认当前自然年"),
            ("任务模式", "标准 或 快速（可选，默认标准）"),
            (
                "示例",
                "账号 zhangsan 密码 （示例） / zhangsan@example.com / / 备注 / 2026,2025 / 标准"
                if combined
                else "zhangsan@example.com / （密码）/ 备注 / 2026,2025 / 标准",
            ),
            ("重复账号", "重复账号自动跳过，不会覆盖"),
        ]
    else:
        notes = [
            ("说明", "内容"),
            ("表头", "不可改字、不可调列顺序"),
            ("必填列", "账号+密码（分列填写），或仅填「账号密码」一栏（自动识别）"),
            *(
                [("账号密码", "可粘贴「账号 xxx 密码 yyy」或 user pass；与分列二选一")]
                if combined else []
            ),
            ("学科", "学科与学分成对填写；学分支持 0.5；未启用学科需求可留空"),
            ("卡号", "仅站点支持购卡/充值时填写"),
            ("重复账号", "重复账号自动跳过，不会覆盖"),
            ("导入格式", "仅解析「账号列表」Sheet 的中文表头行"),
        ]
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 50
    for ri, (k, v) in enumerate(notes, start=1):
        ws2.cell(row=ri, column=1, value=k).font = Font(name="微软雅黑", bold=True)
        ws2.cell(row=ri, column=2, value=v).font = Font(name="微软雅黑")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _get_sample_row(cols: list[str]) -> list[str]:
    combined = COMBINED_CRED_COL in cols
    mapping = {
        "姓名": "张三",
        COMBINED_CRED_COL: "账号 zhangsan@example.com 密码 （示例密码）",
        "账号": "" if combined else "zhangsan@example.com",
        "密码": "" if combined else "（示例密码）",
        "学科1": "内科学",
        "学分1": "5",
        "学科2": "公共课",
        "学分2": "3",
        "卡号": "（如有）",
        "卡号密码": "（如有）",
        "备注": "备注信息",
        "目标年度": "2026,2025",
        "任务模式": "标准",
    }
    return [mapping.get(c, "") for c in cols]


# ── 导入解析 ──────────────────────────────────────────────────────────────────

@dataclass
class ImportRow:
    display_name: str
    username: str
    password: str
    requirements: list[dict]       # A 型：[{"category":..., "credits":...}]
    target_years: list[str]        # B 型
    report_mode: str
    extra: dict[str, Any]
    remark: str


@dataclass
class ImportResult:
    rows: list[ImportRow]
    added: int
    skipped: int
    failed: int
    errors: list[dict]  # {"row": int, "reason": str}


def parse_import_xlsx(
    data: bytes,
    site_profile: str = "A",
    credential_input_mode: str = "split",
) -> ImportResult:
    """
    解析导入文件。
    site_profile: "A"（学科规划型）或 "B"（公需年度型）
    只认中文表头；英文表头报错。
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if "账号" in sn or sn == wb.sheetnames[0]:
            ws = wb[sn]
            break
    if ws is None:
        return ImportResult([], 0, 0, 1, [{"row": 0, "reason": "未找到账号 Sheet"}])

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return ImportResult([], 0, 0, 1, [{"row": 0, "reason": "Sheet 为空"}])

    header = [str(c).strip() if c else "" for c in header_row]

    english_headers = [h for h in header if h and h.isascii() and h.replace(" ", "").isalpha() and len(h) > 1]
    if len(english_headers) > 3:
        return ImportResult([], 0, 0, 1, [{
            "row": 1,
            "reason": f"表头必须为中文，检测到英文列名：{', '.join(english_headers[:5])}",
        }])

    def col(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    rows: list[ImportRow] = []
    errors: list[dict] = []

    for ri, row in enumerate(rows_iter, start=2):
        def v_at(idx: int | None, default: str = "") -> str:
            if idx is None or row[idx] in (None, ""):
                return default
            return str(row[idx]).strip()

        def v(name: str, default: str = "") -> str:
            return v_at(col(name), default)

        combined_raw = v(COMBINED_CRED_COL) if col(COMBINED_CRED_COL) is not None else ""

        if site_profile == "B":
            username = v("账号")
            password = v("密码")
            if not username and not password and not combined_raw:
                continue
            try:
                username, password = resolve_credentials_from_row(username, password, combined_raw)
            except CredentialParseError as exc:
                errors.append({"row": ri, "reason": str(exc)})
                continue
            remark = v("备注")
            yi = _resolve_header_col(header, B_YEAR_HEADER_ALIASES)
            years_raw = v_at(yi)
            target_years = (
                [y.strip() for y in years_raw.replace("，", ",").replace("、", ",").split(",") if y.strip()]
                if years_raw else [_current_year_str()]
            )
            mi = _resolve_header_col(header, B_MODE_HEADER_ALIASES)
            report_mode = v_at(mi, "normal")
            if report_mode in ("快速", "fast"):
                report_mode = "fast"
            else:
                report_mode = "normal"
            rows.append(ImportRow(
                display_name="",
                username=username,
                password=password,
                requirements=[],
                target_years=target_years,
                report_mode=report_mode,
                extra={"remark": remark},
                remark=remark,
            ))
        else:
            username = v("账号")
            password = v("密码")
            if not username and not password and not combined_raw:
                continue
            try:
                username, password = resolve_credentials_from_row(username, password, combined_raw)
            except CredentialParseError as exc:
                errors.append({"row": ri, "reason": str(exc)})
                continue
            display_name = v("姓名", "")
            remark = v("备注")
            card_no = v("卡号")
            card_pwd = v("卡号密码")
            reqs = []
            for i in range(1, 5):
                cat = v(f"学科{i}")
                cr_raw = v(f"学分{i}", "")
                if cat and not cr_raw:
                    errors.append({"row": ri, "reason": f"学科{i} 已填但学分{i} 为空"})
                    break
                if cr_raw and not cat:
                    errors.append({"row": ri, "reason": f"学分{i} 已填但学科{i} 为空"})
                    break
                try:
                    cr = float(cr_raw) if cr_raw else 0.0
                except ValueError:
                    errors.append({"row": ri, "reason": f"学分{i} 不是有效数字"})
                    break
                if cat:
                    reqs.append({"category": cat, "credits": cr})
            else:
                extra: dict[str, Any] = {}
                if card_no:
                    extra["card_no"] = card_no
                if card_pwd:
                    extra["card_password"] = card_pwd
                if remark:
                    extra["remark"] = remark
                rows.append(ImportRow(
                    display_name=display_name or username,
                    username=username,
                    password=password,
                    requirements=reqs,
                    target_years=[],
                    report_mode="normal",
                    extra=extra,
                    remark=remark,
                ))

    return ImportResult(rows=rows, added=0, skipped=0, failed=len(errors), errors=errors)


# ── 导出 ──────────────────────────────────────────────────────────────────────

def build_export_xlsx(
    accounts: list[dict],
    site_profile: str = "A",
    import_cols: list[str] | None = None,
    credential_input_mode: str = "split",
) -> bytes:
    """
    导出：前 N 列与导入模板完全一致；后追加状态/日志等系统列。
    accounts: list of dict（须含 username；combined 模式须含 password 以填充「账号密码」列）
    """
    cols = import_cols or import_cols_for(site_profile, credential_input_mode)
    extra_cols = EXPORT_EXTRA_COLS if site_profile.upper() == "A" else B_EXPORT_EXTRA_COLS
    all_cols = cols + extra_cols

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "账号列表"
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill("solid", fgColor="3659F0")
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for ci, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col_name) * 2 + 4)

    status_map = {
        "queued": "排队", "running": "进行中", "waiting_apply": "等待申请",
        "completed": "已完成", "failed": "失败", "paused": "已暂停",
    }

    for ri, acc in enumerate(accounts, start=2):
        extra = json.loads(acc.get("extra_json") or "{}")
        reqs = json.loads(acc.get("requirements_json") or "[]")
        target_years = json.loads(acc.get("target_years_json") or "[]")

        username = acc.get("username", "") or ""
        password_plain = acc.get("password", "") or ""
        row_data: dict[str, Any] = {
            "姓名": acc.get("display_name", "") or extra.get("real_name", ""),
            "身份证": extra.get("id_card", ""),
            "账号": username,
            "密码": "",
            COMBINED_CRED_COL: (
                format_combined_credential(username, password_plain)
                if COMBINED_CRED_COL in cols
                else ""
            ),
            "备注": extra.get("remark", ""),
            "卡号": extra.get("card_no", ""),
            "卡号密码": "",
            "目标年度": "、".join(target_years),
            "任务模式": "快速" if extra.get("report_mode") == "fast" else "标准",
            "状态": status_map.get(acc.get("status", ""), acc.get("status", "")),
            "说明": acc.get("status_msg", ""),
            "重试次数": acc.get("retry_count", 0),
            "创建时间": _fmt_ts(acc.get("created_at")),
            "更新时间": _fmt_ts(acc.get("updated_at")),
            "最近运行结果": acc.get("last_run_result", ""),
            "错误日志": acc.get("error_log_text", ""),
        }

        # A 型：学科/学分列
        for i, req in enumerate(reqs[:4], start=1):
            row_data[f"学科{i}"] = req.get("category", "")
            row_data[f"学分{i}"] = req.get("credits", "")

        for ci, col_name in enumerate(all_cols, start=1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt_ts(ts: float | None) -> str:
    if not ts:
        return ""
    import datetime
    return datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
